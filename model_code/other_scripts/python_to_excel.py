import pandas as pd
import openpyxl

def python_to_excel(output_dir, wb_filename, small_area):

  #read in the template
  if(small_area == "ward"):
    book = openpyxl.load_workbook('input_data/excel_templates/ward_housing_led_2020_based_template.xlsx')
  elif(small_area == "msoa"):
    book = openpyxl.load_workbook('input_data/excel_templates/msoa_housing_led_2020_based_template.xlsx')
  else:
    return("Error in python_excel_output: small_area must be either ward or msoa")

  #read in the data from csv
  persons = pd.read_csv(output_dir+small_area+"/persons_"+small_area+".csv")
  females = pd.read_csv(output_dir+small_area+"/females_"+small_area+".csv")
  males = pd.read_csv(output_dir+small_area+"/males_"+small_area+".csv")
  components = pd.read_csv(output_dir+small_area+"/components_"+small_area+".csv")
  
  #change underscore to space in column headers
  persons.columns = persons.columns.str.replace('_',' ')
  females.columns = females.columns.str.replace('_',' ')
  males.columns = males.columns.str.replace('_',' ')
  components.columns = components.columns.str.replace('_',' ')

  #Set the output name and prep the file for writing
  writer = pd.ExcelWriter(output_dir+"excels/"+wb_filename, engine='openpyxl') 
  writer.book = book
  writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

  #Write the dataframes to the workbook
  persons.to_excel(writer, "persons", index=False)
  females.to_excel(writer, "females", index=False)
  males.to_excel(writer, "males", index=False)
  components.to_excel(writer, "components of change", index=False)
  
  # Format the header row - no border, left aligned
  side = openpyxl.styles.Side(border_style=None)
  no_border = openpyxl.styles.borders.Border(
    left=side, right=side, top=side, bottom=side,
  )

  for ws in book.worksheets:
    for cell in ws["1:1"]:
      cell.border = no_border
      cell.alignment = openpyxl.styles.Alignment(horizontal='left')
  
  # Save the workbook
  writer.save()
