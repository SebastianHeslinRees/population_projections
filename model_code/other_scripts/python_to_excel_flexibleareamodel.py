import pandas as pd
import openpyxl
import datetime

def python_to_excel_flexibleareamodel(persons, females, males, components, trajectory, stock, output_dir, wb_filename, projection_name, small_area):

  #read in the template
  if(small_area == "ward"):
    book = openpyxl.load_workbook('input_data/excel_templates/ward_housing_led_2021_based_template.xlsx')
  elif(small_area == "msoa"):
    book = openpyxl.load_workbook('input_data/excel_templates/msoa_housing_led_2021_based_template.xlsx')
  else:
    return("Error in python_excel_output: small_area must be either ward or msoa")
  
  #change underscore to space in column headers
  persons.columns = persons.columns.str.replace('_',' ')
  females.columns = females.columns.str.replace('_',' ')
  males.columns = males.columns.str.replace('_',' ')
  components.columns = components.columns.str.replace('_',' ')
  trajectory.columns = trajectory.columns.str.replace('_',' ')
  stock.columns = stock.columns.str.replace('_',' ')

  #Set the output name and prep the file for writing
  writer = pd.ExcelWriter(output_dir+wb_filename, engine='openpyxl') 
  writer.book = book
  writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

  #Write the dataframes to the workbook
  persons.to_excel(writer, "persons", index=False)
  females.to_excel(writer, "females", index=False)
  males.to_excel(writer, "males", index=False)
  components.to_excel(writer, "components of change", index=False)
  trajectory.to_excel(writer, "dwelling trajectory", index=False)
  stock.to_excel(writer, "dwelling stock", index=False)
  
  book['Metadata']['A3'] = projection_name
  book['Metadata']['A8'] = 'Projection produced on ' + datetime.date.today().strftime("%d %B, %Y")
   
  #book.get_sheet_by_name('Metadata')['A3'] = projection_name
  #book.get_sheet_by_name('Metadata')['A8'] = 'Projection produced on ' + datetime.date.today().strftime("%d %B, %Y")
  
   
  #Remove border formatting that openpyxl adds
  side = openpyxl.styles.Side(border_style=None)
  no_border = openpyxl.styles.borders.Border(
    left=side, right=side, top=side, bottom=side,
  )

  #Left align header row
  for ws in book.worksheets:
    for cell in ws["1:1"]:
      cell.border = no_border
      cell.alignment = openpyxl.styles.Alignment(horizontal='left')
      
  #Set Excel number format
  for ws in book.worksheets:
    if ws.title != "Metadata":
      for row_cells in ws.iter_rows(min_col=3, max_col=44):
        for cell in row_cells:
          cell.number_format = '0'
  
  # Save the workbook
  writer.save()
