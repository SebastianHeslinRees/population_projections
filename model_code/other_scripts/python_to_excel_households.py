import pandas as pd
import openpyxl

def python_to_excel_households(output_dir, wb_filename, projection_name, model):

  #read in the template
  book = openpyxl.load_workbook('input_data/excel_templates/household_template_2020.xlsx')
  hh_output_dir = output_dir+"households/"+model+"_"
  
  #read in the data from csv
  stage1_households = pd.read_csv(hh_output_dir+"stage1_households.csv")
  stage2_households = pd.read_csv(hh_output_dir+"stage2_households.csv")
  detailed_ce_pop = pd.read_csv(hh_output_dir+"detailed_ce_pop.csv")
  detailed_hh_pop = pd.read_csv(hh_output_dir+"detailed_hh_pop.csv")
  household_summary = pd.read_csv(hh_output_dir+"household_summary.csv") 

  #change underscore to space in column headers
  stage1_households.columns = stage1_households.columns.str.replace('_',' ')
  stage2_households.columns = stage2_households.columns.str.replace('_',' ')
  detailed_ce_pop.columns = detailed_ce_pop.columns.str.replace('_',' ')
  detailed_hh_pop.columns = detailed_hh_pop.columns.str.replace('_',' ')
  household_summary.columns = household_summary.columns.str.replace('_',' ')

  #Set the output name and prep the file for writing
  writer = pd.ExcelWriter(output_dir+"datastore/"+wb_filename, engine='openpyxl') 
  writer.book = book
  writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

  #Write the dataframes to the workbook
  stage1_households.to_excel(writer, "stage 1 households", index=False)
  stage2_households.to_excel(writer, "stage 2 households", index=False)
  detailed_ce_pop.to_excel(writer, "household popn", index=False)
  detailed_hh_pop.to_excel(writer, "communal est popn", index=False)
  household_summary.to_excel(writer, "summary", index=False)
  
  book.get_sheet_by_name('Metadata')['A3'] = projection_name+" - "+model.upper()+" houshold model"
  
  #Remove border formatting that openpyxl adds
  side = openpyxl.styles.Side(border_style=None)
  no_border = openpyxl.styles.borders.Border(
    left=side, right=side, top=side, bottom=side,
  )

  #Left alight header row
  for ws in book.worksheets:
    for cell in ws["1:1"]:
      cell.border = no_border
      cell.alignment = openpyxl.styles.Alignment(horizontal='left')
      
  #Set Excel number format
  for ws in book.worksheets:
    if ws.title != "Metadata" and ws.title != "summary":
      for row_cells in ws.iter_rows(min_col=3, max_col=44):
        for cell in row_cells:
          cell.number_format = '0'
          
  ws = book['summary']
  for row_cells in ws.iter_rows(min_col=3, max_col=6):
    for cell in row_cells:
      cell.number_format = '0'
  
  # Save the workbook
  writer.save()
