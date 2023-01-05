import pandas as pd
import openpyxl

def python_to_excel_trendmodel(persons, females, males, components, output_dir, wb_filename, projection_name):

  #read in the template
  book = openpyxl.load_workbook('input_data/excel_templates/trend_template_2021.xlsx')
  
  #Set the output name and prep the file for writing
  writer = pd.ExcelWriter(output_dir+"/"+wb_filename, engine='openpyxl') 
  writer.book = book
  writer.worksheets = dict((ws.title, ws) for ws in book.worksheets)

  #Write the dataframes to the workbook
  persons.to_excel(writer, "persons", index=False)
  females.to_excel(writer, "females", index=False)
  males.to_excel(writer, "males", index=False)
  components.to_excel(writer, "components of change", index=False)
  
  book['Metadata']['A3']  = projection_name

  #Remove border formatting that openpyxl adds
  side = openpyxl.styles.Side(border_style=None)
  no_border = openpyxl.styles.borders.Border(
    left=side, right=side, top=side, bottom=side,
  )

  #Left alight header row
  for ws in book.worksheets:
    for cell in ws["1:1"]:
      cell.border = no_border
      cell.alignment = openpyxl.styles.Alignment(horizontal='left')
  
  #Set Excel number format
  for ws in book.worksheets:
    if ws.title != "Metadata":
      for row_cells in ws.iter_rows(min_col=3, max_col=44):
        for cell in row_cells:
          cell.number_format = '0'
  
  # Save the workbook
  writer.close()
